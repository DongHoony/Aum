# -*- coding: utf-8 -*-
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import QPushButton, QWidget

class Aum(QWidget):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.setFixedSize(700, 210)

        self.placeButton = QPushButton(Form)
        self.placeButton.setGeometry(QtCore.QRect(490, 180, 75, 23))
        self.placeButton.setObjectName("placeButton")
        self.placeButton.clicked.connect(self.buildList)

        self.exitButton = QtWidgets.QPushButton(Form)
        self.exitButton.setGeometry(QtCore.QRect(600, 180, 75, 23))
        self.exitButton.setObjectName("exitButton")
        self.exitButton.clicked.connect(self.openFile)
        self.exitButton.setEnabled(False)

        self.groupnumSpinbox = QtWidgets.QSpinBox(Form)
        self.groupnumSpinbox.setGeometry(QtCore.QRect(630, 20, 42, 22))
        self.groupnumSpinbox.setObjectName("groupnumSpinbox")

        self.textEdit = QtWidgets.QTextEdit(Form)
        self.textEdit.setGeometry(QtCore.QRect(490, 80, 181, 31))
        self.textEdit.setObjectName("textEdit")

        self.label = QtWidgets.QLabel(Form)
        self.label.setGeometry(QtCore.QRect(580, 20, 56, 21))
        self.label.setObjectName("label")

        self.label_2 = QtWidgets.QLabel(Form)
        self.label_2.setGeometry(QtCore.QRect(490, 55, 181, 21))
        self.label_2.setObjectName("label_2")

        self.textBrowser = QtWidgets.QTextBrowser(Form)
        self.textBrowser.setGeometry(QtCore.QRect(20, 10, 421, 193))
        self.textBrowser.setObjectName("textBrowser")
        self.textBrowser.append("====== 아름다운배움 꿈사다리학교 멘토 임시배치 프로그램 ======")
        self.textBrowser.append("같은 폴더에 aumpeople.xlsx 파일 존재 여부를 확인 후 배치해주세요.")

        self.progressBar = QtWidgets.QProgressBar(Form)
        self.progressBar.setGeometry(QtCore.QRect(490, 130, 181, 23))
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")

        self.retranslateUi(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "꿈사다리학교 멘토배치 프로그램"))
        self.placeButton.setText(_translate("Form", "배치"))
        self.exitButton.setText(_translate("Form", "파일 열기"))
        self.label.setText(_translate("Form", "그룹 수"))
        self.label_2.setText(_translate("Form", "그룹별 정원 수, \'/\'로 구분"))

    def buildList(self):
        global name, ALL_MAN_RATIO
        import openpyxl
        from collections import Counter

        while 1:
            text = self.textEdit.toPlainText()
            try:
                r_max_people = list(map(int, text.split('/')))
            except ValueError:
                self.textBrowser.append("잘못된 형식입니다. 다시 입력해 주세요.")
                return 0
            numberofgroup = self.groupnumSpinbox.value()
            print(numberofgroup)
            if len(r_max_people) != int(numberofgroup) or int(numberofgroup) == 0:
                self.textBrowser.append("그룹 수와 그룹별 인원이 일치하지 않습니다. 다시 입력해 주세요.")
                return 0
            else:
                break
        try:
            f = openpyxl.load_workbook(r"aumpeople.xlsx")
            sheet = f.active
        except FileNotFoundError:
            self.textBrowser.append("파일이 존재하지 않습니다. 확인 후 배치해 주세요.")
            return 0

        # 성비 확인, list형태로 내보낸다, list 안에는 tuple 형태로 (W : M).
        def RatioCheck():
            ratio = []
            for i in range(NUMBERS_OF_REGIONS):
                s = list(map(list, zip(*regions[i])))[1]
                w = s.count('여자')
                m = s.count('남자')
                ratio.append((round(w * 100 / (w + m), 1), round(m * 100 / (w + m), 1)))
            return ratio

        # 정원 확인, 넘치면 양수, 모자라면 음수.
        def PersonCheck():
            ppl = []
            for i in range(NUMBERS_OF_REGIONS):
                ppl.append(len(regions[i]) - r_max_people[i])
            return ppl

        # 남자 성비 확인
        def GetManRatio():
            s = list(map(list, zip(*regions[i])))[1]
            m = s.count('남자')
            w = s.count('여자')
            manratio = round(m * 100 / (w + m), 1)
            return manratio

        REGIONS_ALPHABET = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        name = []

        # 모든 명단 리스트로 작성
        for r in sheet.rows:
            s = []
            # s => [이름, 성별, 나이, 1지망, 2지망, 연락처, 기존/신규]
            for i in range(1, 8):
                s.append(r[i].value)
            name.append(s)

        # 서식 맨 위 구분 삭제
        name.pop(0)

        # 전체 성비 구하기
        genders = list(map(list, zip(*name)))[1]
        w = genders.count('여자')
        m = genders.count('남자')
        ALL_MAN_RATIO = round(m * 100 / (w + m), 1)

        # 각 배열에 그룹 입력
        NUMBERS_OF_REGIONS = int(numberofgroup)
        regions = [[] for i in range(NUMBERS_OF_REGIONS)]
        temp = []
        reg = REGIONS_ALPHABET[:NUMBERS_OF_REGIONS]
        for i in name:
            try:
                regions[reg.index(i[3])].append(i)
            except ValueError:
                self.textBrowser.append("입력하신 그룹 수와 엑셀 내 그룹 수와 일치하지 않습니다. 확인해 주세요.")
                return 0

        self.placeButton.setEnabled(False)
        self.textEdit.setEnabled(False)
        self.groupnumSpinbox.setEnabled(False)
        self.exitButton.setEnabled(False)

        _ = 0
        while 1:  # While
            self.progressBar.setProperty("value", _)
            _ += 1
            self.textBrowser.append("ATTEMPT {}".format(_))
            ppl = PersonCheck()
            isChanged = 0
            for i in range(NUMBERS_OF_REGIONS):
                g = reg[i]
                if ppl[i] == 0:
                    self.textBrowser.append('OK with {}, continue...'.format(g))
                elif ppl[i] > 0:
                    self.textBrowser.append('Too many people in {}, pushing people to temporary list...'.format(g))
                    for j in range(ppl[i]):
                        for a in regions[i]:
                            cnt = Counter(list(map(list, zip(*regions[i])))[2])
                            most_age = cnt.most_common(1)[0][0]
                            manratio = GetManRatio()
                            if manratio < ALL_MAN_RATIO:
                                if a[1] == '여자' and a[2] == most_age and a[6] == '신규':
                                    self.textBrowser.append("Pushing {} into temporary list...".format(a[0]))
                                    temp.append(a)
                                    regions[i].pop(regions[i].index(a))
                                    isChanged = 1
                                    break
                            else:
                                if a[1] == '남자' and a[2] == most_age and a[6] == '신규':
                                    self.textBrowser.append("Pushing {} into temporary list...".format(a[0]))
                                    temp.append(a)
                                    regions[i].pop(regions[i].index(a))
                                    isChanged = 1
                                    break
                    if not isChanged:
                        ppl = PersonCheck()
                        self.textBrowser.append('Nobody matches to pull out, making the gender ratio ideally...')
                        for j in range(ppl[i]):
                            for a in regions[i]:
                                manratio = GetManRatio()
                                if manratio < ALL_MAN_RATIO:
                                    if a[1] == '여자' and a[6] == '신규':
                                        self.textBrowser.append("Pushing {} into temporary list...".format(a[0]))
                                        temp.append(a)
                                        regions[i].pop(regions[i].index(a))
                                        isChanged = 1
                                        break
                                else:
                                    if a[1] == '남자' and a[6] == '신규':
                                        self.textBrowser.append("Pushing {} into temporary list...".format(a[0]))
                                        temp.append(a)
                                        regions[i].pop(regions[i].index(a))
                                        isChanged = 1
                                        break
                    if not isChanged:
                        self.textBrowser.append("Members are all old, pulling old members...")
                        ppl = PersonCheck()
                        for j in range(ppl[i]):
                            for a in regions[i]:
                                manratio = GetManRatio()
                                if manratio < ALL_MAN_RATIO:
                                    if a[1] == '여자' and a[6] == '기존':
                                        self.textBrowser.append("Pushing {} into temporary list...".format(a[0]))
                                        temp.append(a)
                                        regions[i].pop(regions[i].index(a))
                                        isChanged = 1
                                        break
                                else:
                                    if a[1] == '남자' and a[6] == '기존':
                                        self.textBrowser.append("Pushing {} into temporary list...".format(a[0]))
                                        temp.append(a)
                                        regions[i].pop(regions[i].index(a))
                                        isChanged = 1
                                        break
                else:
                    self.textBrowser.append('Low people in {}, pulling people from temporary list...'.format(g))
                    if not temp:
                        self.textBrowser.append('No people in temporary list, continue...'.format(g))
                    else:
                        for a in temp:
                            manratio = GetManRatio()
                            if manratio < ALL_MAN_RATIO:
                                if a[1] == '남자' and a[4] == reg[i]:
                                    self.textBrowser.append("Pulling {} from temporary list...".format(a[0]))
                                    regions[i].append(temp.pop(temp.index(a)))
                                    isChanged = 1
                            else:
                                if a[1] == '여자' and a[4] == reg[i]:
                                    self.textBrowser.append("Pulling {} from temporary list...".format(a[0]))
                                    regions[i].append(temp.pop(temp.index(a)))
                                    isChanged = 1
                    if not isChanged:
                        self.textBrowser.append('Nobody matches on temporary list, continue...')
            if not isChanged or _ > 100:
                break
        self.progressBar.setProperty("value", 90)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '임시명단'
        ws.merge_cells('A1:G1')
        for i in range(7):
            ws.cell(row=2, column=i + 1).value = ['이름', '성별', '나이', '1지망', '2지망', '연락처', '기존/신규'][i]
        ws.cell(row=1, column=1).value = '임시 명단 : {}명'.format(len(temp))
        for r in range(len(temp)):
            for k in range(len(temp[r])):
                ws.cell(row=r + 3, column=k + 1).value = temp[r][k]

        for i in range(NUMBERS_OF_REGIONS):
            ws = wb.create_sheet(reg[i])
            ws.merge_cells('A1:G1')
            for _ in range(7):
                ws.cell(row=2, column=_ + 1).value = ['이름', '성별', '나이', '1지망', '2지망', '연락처', '기존/신규'][_]
            ws.cell(row=1, column=1).value = "{} ({} / {})".format(reg[i], len(regions[i]), r_max_people[i])
            for r in range(len(regions[i])):
                for k in range(len(regions[i][r])):
                    ws.cell(row=r + 3, column=k + 1).value = regions[i][r][k]
        self.progressBar.setProperty("value", 100)

        wb.save(r'aumpeople_sorted.xlsx')
        self.exitButton.setEnabled(True)
        self.textBrowser.append("\n" * 20 )
        self.textBrowser.append("\n총 인원 : {}명, 전체 성비 -> W {} : M {}\n".format(len(name), 100 - ALL_MAN_RATIO, ALL_MAN_RATIO))
        self.textBrowser.append("======" * 10)
        self.textBrowser.append("자동 배치가 완료되었습니다.\n"
              "해당 배치는 엑셀 순번에 따라 정원수를 기준으로 작성되었습니다. \n최종 합격자 명단은 아닙니다.\n"
              "임시 정원외 명단으로 배치된 경우는 :\n1. 그룹의 정원 초과\n2. 그룹 내 성비 조정\n"
              "으로 자동 조정된 것이며, 자동배치이므로 임시 정원외 명단도 확인해 주세요.\n" + "======" * 10)
        self.textBrowser.append("같은 폴더에 aumpeople_sorted.xlsx로 저장되었습니다.")
        self.exitButton.setEnabled(True)
        return 0

    def openFile(self):
        import os
        os.startfile('aumpeople_sorted.xlsx')
        return 0


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = Aum()
    ui.setupUi(Form)
    Form.show()
    sys.exit(app.exec_())
