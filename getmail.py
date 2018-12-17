from selenium import webdriver
import openpyxl
from time import sleep
from math import ceil
from bs4 import BeautifulSoup

driver = webdriver.Chrome('chromedriver.exe')
driver.get('https://nid.naver.com/nidlogin.login')
input("Enter if you're ready")
driver.get("https://mail.naver.com")
file = openpyxl.load_workbook('mail_sort.xlsx')
sheet = file.active
howmany = ceil(int(input("How many pages?"))/10)
try:
    for pages in range(howmany):
        for buttons in range(10):
            sleep(0.3)
            html = driver.page_source
            soup = BeautifulSoup(html, 'lxml')
            title_list = soup.find_all('strong', 'mail_title')
            from_list = soup.find_all('div', 'mTitle')
            date = soup.find_all('li', 'iDate')
            for i in range(30):
                print(str(from_list[i].find('a').get('title')).replace("\"", "", 2).replace(">","").replace(" <", ": "))
                print(title_list[i].text)
                sheet.cell(row=pages*300+buttons*30+i+1, column=1).value=str(from_list[i].find('a').get('title')).replace("\"", "", 2).replace(">","").replace(" <", ": ")
                sheet.cell(row=pages*300+buttons*30+i+1, column=2).value=title_list[i].text
                sheet.cell(row=pages * 300 + buttons * 30 + i + 1, column=3).value = date[i].text
            if buttons != 9:
                driver.find_element_by_xpath('//*[@id="{}"]'.format(pages*10+(buttons+2))).click()
                sleep(0.5)
        driver.find_element_by_xpath('//*[@id="next_page"]').click()
        sleep(0.5)
except IndexError:
    file.save('mail_sorted.xlsx')
    exit()
file.save('mail_sorted.xlsx')
exit()
