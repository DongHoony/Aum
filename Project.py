import openpyxl
from openpyxl.styles import Alignment
from collections import Counter
print("======"*10+"\n아름다운배움 꿈사다리학교 멘토 임시배치 프로그램입니다.\n\n")
input("프로그램이 있는 폴더에 aumpeople.xlsx라는 파일이 있는지 확인 후, 엔터를 입력해 주세요.")
try:
    f = openpyxl.load_workbook(r"aumpeople.xlsx")
except FileNotFoundError:
    if input("파일이 존재하지 않습니다. 엔터를 입력하면 종료합니다."):
        exit()
sheet = f.active

# 성비 확인, list형태로 내보낸다, list 안에는 tuple 형태로 (W : M).
def RatioCheck():
    ratio = []
    for i in range(NUMBERS_OF_REGIONS):
        s = list(map(list, zip(*regions[i])))[1]
        w = s.count('여자')
        m = s.count('남자')
        ratio.append((round(w * 100 / (w + m), 1), round(m * 100 / (w + m), 1)))
    return ratio

# 정원 확인, 넘치면 양수, 모자라면 음수.
def PersonCheck():
    ppl = []
    for i in range(NUMBERS_OF_REGIONS):
        ppl.append(len(regions[i]) - r_max_people[i])
    return ppl

#남자 성비 확인
def GetManRatio():
    s = list(map(list, zip(*regions[i])))[1]
    m = s.count('남자')
    w = s.count('여자')
    manratio = round(m * 100 / (w + m), 1)
    return manratio


REGIONS_ALPHABET = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
name = []

# 모든 명단 리스트로 작성
for r in sheet.rows:
    s = []
    # s => [이름, 성별, 나이, 1지망, 2지망, 연락처, 기존/신규]
    for i in range(1, 8):
        s.append(r[i].value)
    name.append(s)

#서식 맨 위 구분 삭제
name.pop(0)

# 전체 성비 구하기
genders = list(map(list, zip(*name)))[1]
w = genders.count('여자')
m = genders.count('남자')
ALL_MAN_RATIO = round(m * 100 / (w + m), 1)
print("\n총 인원 : {}명, 전체 성비 -> W {} : M {}\n".format(len(name), 100-ALL_MAN_RATIO, ALL_MAN_RATIO))

# 각 배열에 그룹 입력
NUMBERS_OF_REGIONS = int(input("그룹 수를 입력해주세요.\n"))
regions = [[] for i in range(NUMBERS_OF_REGIONS)]
while 1:
    r_max_people = list(map(int, input("\n각 그룹의 정원수를 입력해주세요. 정원수는 슬래시'/' 로 구분합니다. \n(예: 10/20/30/20/10)\n").split('/')))
    if len(r_max_people) != NUMBERS_OF_REGIONS:
        print("입력하신 그룹 수에 해당하는 정원 수가 입력되지 않았습니다.\n")
    else:
        break
temp = []
reg = REGIONS_ALPHABET[:NUMBERS_OF_REGIONS]
for i in name:
    regions[reg.index(i[3])].append(i)

_ = 0
while 1: # While
    _ += 1
    print("ATTEMPT {}".format(_))
    ppl = PersonCheck()
    isChanged = 0
    for i in range(NUMBERS_OF_REGIONS):
        g = reg[i]
        if ppl[i] == 0:
            print('OK with {}, continue...'.format(g))

        elif ppl[i] > 0:
            print('Too many people in {}, pushing people to temporary list...'.format(g))
            for j in range(ppl[i]):
                for a in regions[i]:
                    cnt = Counter(list(map(list, zip(*regions[i])))[2])
                    most_age = cnt.most_common(1)[0][0]
                    manratio = GetManRatio()
                    if manratio < ALL_MAN_RATIO:
                        if a[1] == '여자' and a[2] == most_age and a[6] == '신규':
                            print("Pushing {} into temp...".format(a[0]))
                            temp.append(a)
                            regions[i].pop(regions[i].index(a))
                            isChanged = 1
                            break
                    else:
                        if a[1] == '남자' and a[2] == most_age and a[6] == '신규':
                            print("Pushing {} into temp...".format(a[0]))
                            temp.append(a)
                            regions[i].pop(regions[i].index(a))
                            isChanged = 1
                            break
            if not isChanged:
                print('Nobody matches to pull out, making the gender ratio ideally...')
                for j in range(ppl[i]):
                    for a in regions[i]:
                        manratio = GetManRatio()
                        if manratio < ALL_MAN_RATIO:
                            if a[1] == '여자'  and a[6] == '신규':
                                print("Pushing {} into temp...".format(a[0]))
                                temp.append(a)
                                regions[i].pop(regions[i].index(a))
                                isChanged = 1
                                break
                        else:
                            if a[1] == '남자' and a[6] == '신규':
                                print("Pushing {} into temp...".format(a[0]))
                                temp.append(a)
                                regions[i].pop(regions[i].index(a))
                                isChanged = 1
                                break
        else:
            print('Low people in {}, pulling people from temporary list...'.format(g))
            if not temp:
                print('No people in temporary list, continue...'.format(g))
            else:
                for a in temp:
                    manratio = GetManRatio()
                    if manratio < ALL_MAN_RATIO :
                        if a[1] == '남자' and a[4] == reg[i]:
                            print("Pulling {} from temp...".format(a[0]))
                            regions[i].append(temp.pop(temp.index(a)))
                            isChanged = 1
                    else:
                        if a[1] == '여자' and a[4] == reg[i]:
                            print("Pulling {} from temp...".format(a[0]))
                            regions[i].append(temp.pop(temp.index(a)))
                            isChanged = 1
            if not isChanged:
                print('Nobody matches on temporary list, continue...')
    if not isChanged or _ > 100:
        break

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '임시명단'
ws.merge_cells('A1:G1')
for i in range(7):
    ws.cell(row=2,column=i+1).value = ['이름', '성별', '나이', '1지망', '2지망', '연락처', '기존/신규'][i]
ws.cell(row=1, column=1).value = '임시 명단 : {}명'.format(len(temp))
for r in range(len(temp)):
    for k in range(len(temp[r])):
        ws.cell(row=r+3, column=k+1).value = temp[r][k]

for i in range(NUMBERS_OF_REGIONS):
    ws = wb.create_sheet(reg[i])
    ws.merge_cells('A1:G1')
    for _ in range(7):
        ws.cell(row=2, column=_ + 1).value = ['이름', '성별', '나이', '1지망', '2지망', '연락처', '기존/신규'][_]
    ws.cell(row=1, column=1).value = "{} ({} / {})".format(reg[i], len(regions[i]), r_max_people[i])
    for r in range(len(regions[i])):
        for k in range(len(regions[i][r])):
            ws.cell(row=r+3, column=k+1).value = regions[i][r][k]

wb.save(r'aumpeople_sorted.xlsx')
print("\n"*20+"======"*10)
print("자동 배치가 완료되었습니다.\n\n"
      "해당 배치는 엑셀 순번에 따라 정원수를 기준으로 작성되었습니다. \n\n최종 합격자 명단은 아닙니다.\n\n"
      "임시 정원외 명단으로 배치된 경우는 :\n\n1. 그룹의 정원 초과\n2. 그룹 내 성비 조정\n\n"
      "으로 자동 조정된 것이며, 자동배치이므로 임시 정원외 명단도 확인해 주세요.\n\n"+"======"*10)
if input("같은 폴더에 aumpeople_sorted.xlsx로 저장되었습니다. 엔터키를 눌러 종료하십시오."):
    exit()
